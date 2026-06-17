from flask import render_template, redirect, url_for, flash, request, Blueprint, session
from app import db, bcrypt
from app.models import User, Product, Movement, Order, Notification
from app.forms import RegistrationForm, LoginForm, ProductForm, MovementForm
from flask_login import login_user, current_user, logout_user, login_required
import secrets
from datetime import datetime

bp = Blueprint('routes', __name__)


# ==================== ГЛАВНАЯ ====================
@bp.route('/')
@bp.route('/index')
def index():
    products = Product.query.all()
    movements = Movement.query.order_by(Movement.date.desc()).limit(10).all()
    return render_template('index.html', products=products, movements=movements)


# ==================== АВТОРИЗАЦИЯ ====================
@bp.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('routes.index'))
    
    form = RegistrationForm()
    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        user = User(
            username=form.username.data,
            email=form.email.data,
            password=hashed_password,
            role='client'
        )
        db.session.add(user)
        db.session.commit()
        flash('Регистрация успешна!', 'success')
        return redirect(url_for('routes.login'))
    
    return render_template('register.html', form=form)


@bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('routes.index'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and bcrypt.check_password_hash(user.password, form.password.data):
            login_user(user)
            flash(f'Добро пожаловать, {user.username}!', 'success')
            return redirect(url_for('routes.index'))
        else:
            flash('Неверный email или пароль', 'danger')
    
    return render_template('login.html', form=form)


@bp.route('/logout')
def logout():
    logout_user()
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('routes.index'))


# ==================== ПРОФИЛЬ ====================
@bp.route('/profile')
@login_required
def profile():
    user_movements = Movement.query.filter_by(user_id=current_user.id).order_by(Movement.date.desc()).all()
    return render_template('profile.html', movements=user_movements)


# ==================== ТОВАРЫ ====================
@bp.route('/products')
def products():
    search_query = request.args.get('search', '')
    
    if search_query:
        products = Product.query.filter(
            (Product.name.contains(search_query)) | 
            (Product.article.contains(search_query))
        ).all()
    else:
        products = Product.query.all()
    
    return render_template('products.html', products=products, search_query=search_query)


@bp.route('/product/add', methods=['GET', 'POST'])
@login_required
def add_product():
    if not current_user.can_add_product():
        flash('Доступ запрещен. Только для кладовщиков и администраторов', 'danger')
        return redirect(url_for('routes.index'))
    
    form = ProductForm()
    if form.validate_on_submit():
        product = Product(
            name=form.name.data,
            article=form.article.data,
            quantity=form.quantity.data,
            price=form.price.data,
            category=form.category.data,
            unit=form.unit.data,
            manufacturer=form.manufacturer.data,
            min_stock=form.min_stock.data,
            location=form.location.data
        )
        db.session.add(product)
        db.session.commit()
        flash('Товар добавлен', 'success')
        return redirect(url_for('routes.products'))
    
    return render_template('add_product.html', form=form)


@bp.route('/product/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_product(id):
    if not current_user.can_add_product():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.index'))
    
    product = Product.query.get_or_404(id)
    form = ProductForm()
    
    if form.validate_on_submit():
        product.name = form.name.data
        product.article = form.article.data
        product.quantity = form.quantity.data
        product.price = form.price.data
        product.category = form.category.data
        product.unit = form.unit.data
        product.manufacturer = form.manufacturer.data
        product.min_stock = form.min_stock.data
        product.location = form.location.data
        db.session.commit()
        flash('Товар обновлен', 'success')
        return redirect(url_for('routes.products'))
    
    form.name.data = product.name
    form.article.data = product.article
    form.quantity.data = product.quantity
    form.price.data = product.price
    form.category.data = product.category
    form.unit.data = product.unit
    form.manufacturer.data = product.manufacturer
    form.min_stock.data = product.min_stock
    form.location.data = product.location
    
    return render_template('edit_product.html', form=form, product=product)


@bp.route('/product/delete/<int:id>')
@login_required
def delete_product(id):
    if not current_user.is_admin():
        flash('Доступ запрещен. Только для администратора', 'danger')
        return redirect(url_for('routes.index'))
    
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('Товар удален', 'success')
    return redirect(url_for('routes.products'))


# ==================== ДВИЖЕНИЯ ТОВАРОВ ====================
@bp.route('/movements', methods=['GET', 'POST'])
@login_required
def movements():
    form = MovementForm()
    form.product_id.choices = [(p.id, f"{p.name} ({p.quantity} {p.unit})") for p in Product.query.all()]
    
    if form.validate_on_submit():
        product = Product.query.get(form.product_id.data)
        movement = Movement(
            product_id=form.product_id.data,
            user_id=current_user.id,
            type=form.type.data,
            quantity=form.quantity.data,
            comment=form.comment.data
        )
        
        if form.type.data == 'приход':
            product.quantity += form.quantity.data
            flash_msg = f'Приход {form.quantity.data} {product.unit} товара {product.name}'
        elif form.type.data in ['расход', 'продажа']:
            if product.quantity >= form.quantity.data:
                product.quantity -= form.quantity.data
                flash_msg = f'{form.type.data} {form.quantity.data} {product.unit} товара {product.name}'
            else:
                flash(f'Недостаточно товара на складе! В наличии: {product.quantity} {product.unit}', 'danger')
                return redirect(url_for('routes.movements'))
        
        db.session.add(movement)
        db.session.commit()
        flash(flash_msg, 'success')
        return redirect(url_for('routes.index'))
    
    all_movements = Movement.query.order_by(Movement.date.desc()).all()
    return render_template('movements.html', form=form, movements=all_movements)


# ==================== ОТЧЕТЫ ====================
@bp.route('/reports')
@login_required
def reports():
    if not current_user.can_view_reports():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.index'))
    
    products = Product.query.all()
    total_value = sum(p.quantity * p.price for p in products)
    total_quantity = sum(p.quantity for p in products)
    low_stock = [p for p in products if p.quantity <= p.min_stock]
    
    total_movements = Movement.query.count()
    incoming = Movement.query.filter_by(type='приход').count()
    outgoing = Movement.query.filter_by(type='расход').count()
    sales = Movement.query.filter_by(type='продажа').count()
    
    return render_template('reports.html', 
                         products=products,
                         total_quantity=total_quantity,
                         total_value=total_value,
                         low_stock=low_stock,
                         total_movements=total_movements,
                         incoming=incoming,
                         outgoing=outgoing,
                         sales=sales)


@bp.route('/export/report')
@login_required
def export_report():
    """Экспорт отчёта по товарам в Excel"""
    if not current_user.can_view_reports():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.index'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по складу"
    
    headers = ['ID', 'Артикул', 'Наименование', 'Категория', 
               'Производитель', 'Количество', 'Ед. изм.', 'Цена', 'Общая стоимость', 'Место хранения']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    products = Product.query.all()
    for product in products:
        ws.append([
            product.id,
            product.article or '-',
            product.name,
            product.category or '-',
            product.manufacturer or '-',
            product.quantity,
            product.unit,
            product.price,
            product.quantity * product.price,
            product.location or '-'
        ])
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==================== АДМИН-ПАНЕЛЬ ====================
@bp.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.index'))
    
    total_users = User.query.count()
    total_products = Product.query.count()
    total_movements = Movement.query.count()
    total_value = sum(p.quantity * p.price for p in Product.query.all())
    
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()
    recent_movements = Movement.query.order_by(Movement.date.desc()).limit(10).all()
    
    admins = User.query.filter_by(role='admin').count()
    directors = User.query.filter_by(role='director').count()
    storekeepers = User.query.filter_by(role='storekeeper').count()
    managers = User.query.filter_by(role='manager').count()
    accountants = User.query.filter_by(role='accountant').count()
    viewers = User.query.filter_by(role='viewer').count()
    clients = User.query.filter_by(role='client').count()
    
    return render_template('admin_dashboard.html',
                         total_users=total_users,
                         total_products=total_products,
                         total_movements=total_movements,
                         total_value=total_value,
                         recent_users=recent_users,
                         recent_movements=recent_movements,
                         admins=admins,
                         directors=directors,
                         storekeepers=storekeepers,
                         managers=managers,
                         accountants=accountants,
                         viewers=viewers,
                         clients=clients)


@bp.route('/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.index'))
    
    users = User.query.all()
    return render_template('admin_users.html', users=users)


@bp.route('/admin/user/delete/<int:id>')
@login_required
def admin_delete_user(id):
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.index'))
    
    if id == current_user.id:
        flash('Нельзя удалить самого себя', 'danger')
        return redirect(url_for('routes.admin_users'))
    
    user = User.query.get_or_404(id)
    
    movements = Movement.query.filter_by(user_id=id).first()
    if movements:
        flash('Нельзя удалить пользователя с историей операций', 'danger')
        return redirect(url_for('routes.admin_users'))
    
    db.session.delete(user)
    db.session.commit()
    flash(f'Пользователь {user.username} удален', 'success')
    return redirect(url_for('routes.admin_users'))


@bp.route('/admin/change-role/<int:user_id>', methods=['POST'])
@login_required
def admin_change_role(user_id):
    """Изменение роли пользователя (только для администратора)"""
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.index'))
    
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        flash('Нельзя изменить роль своего аккаунта', 'danger')
        return redirect(url_for('routes.admin_users'))
    
    new_role = request.form.get('role')
    allowed_roles = ['admin', 'director', 'storekeeper', 'manager', 'accountant', 'viewer', 'client']
    
    if new_role in allowed_roles:
        user.role = new_role
        db.session.commit()
        flash(f'Роль пользователя {user.username} изменена на "{new_role}"', 'success')
    else:
        flash('Неверная роль', 'danger')
    
    return redirect(url_for('routes.admin_users'))


# ==================== ПОИСК ====================
@bp.route('/search')
def search():
    query = request.args.get('q', '')
    if query:
        products = Product.query.filter(
            (Product.name.contains(query)) | 
            (Product.article.contains(query)) |
            (Product.manufacturer.contains(query))
        ).all()
    else:
        products = []
    
    return render_template('search_results.html', products=products, query=query)


# ==================== ЗАКАЗЫ ====================
@bp.route('/order/<int:product_id>', methods=['GET', 'POST'])
@login_required
def create_order(product_id):
    if current_user.role != 'client':
        flash('Доступ запрещен. Только для клиентов.', 'danger')
        return redirect(url_for('routes.index'))
    
    product = Product.query.get_or_404(product_id)
    
    if request.method == 'POST':
        quantity = float(request.form.get('quantity', 1))
        comment = request.form.get('comment', '')
        
        if quantity <= 0:
            flash('Количество должно быть больше 0', 'danger')
            return redirect(url_for('routes.create_order', product_id=product.id))
        
        if quantity > product.quantity:
            flash(f'Недостаточно товара на складе. В наличии: {product.quantity} {product.unit}', 'danger')
            return redirect(url_for('routes.create_order', product_id=product.id))
        
        order = Order(
            user_id=current_user.id,
            product_id=product.id,
            quantity=quantity,
            comment=comment,
            status='pending'
        )
        db.session.add(order)
        db.session.commit()
        
        flash(f'✅ Заказ на {quantity} {product.unit} товара "{product.name}" успешно оформлен!', 'success')
        return redirect(url_for('routes.my_orders'))
    
    return render_template('create_order.html', product=product)


@bp.route('/my-orders')
@login_required
def my_orders():
    orders = Order.query.filter_by(user_id=current_user.id).order_by(Order.created_at.desc()).all()
    return render_template('my_orders.html', orders=orders)


@bp.route('/admin/orders')
@login_required
def admin_orders():
    if not (current_user.is_admin() or current_user.role == 'manager'):
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.index'))
    
    orders = Order.query.order_by(Order.created_at.desc()).all()
    return render_template('admin_orders.html', orders=orders)


@bp.route('/admin/order/update/<int:order_id>', methods=['POST'])
@login_required
def update_order_status(order_id):
    if not (current_user.is_admin() or current_user.role == 'manager'):
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.index'))
    
    order = Order.query.get_or_404(order_id)
    old_status = order.status
    new_status = request.form.get('status')
    
    status_names = {
        'pending': 'Ожидает',
        'approved': 'Одобрен',
        'rejected': 'Отклонён',
        'completed': 'Выполнен'
    }
    
    if new_status in ['pending', 'approved', 'rejected', 'completed']:
        
        # ========== НОВЫЙ КОД: списание товара при выполнении заказа ==========
        if new_status == 'completed' and old_status != 'completed':
            product = Product.query.get(order.product_id)
            if product.quantity >= order.quantity:
                product.quantity -= order.quantity
                
                # Создаём запись в истории операций
                movement = Movement(
                    product_id=order.product_id,
                    user_id=current_user.id,
                    type='продажа',
                    quantity=order.quantity,
                    comment=f'Заказ #{order.id} выполнен'
                )
                db.session.add(movement)
            else:
                flash(f'Недостаточно товара для выполнения заказа #{order.id}!', 'danger')
                return redirect(url_for('routes.admin_orders'))
        # ====================================================================
        
        order.status = new_status
        db.session.commit()
        
        status_name = status_names.get(new_status, new_status)
        flash(f'Статус заказа #{order.id} изменён на "{status_name}"', 'success')
        
        status_messages = {
            'approved': '✅ Ваш заказ одобрен!',
            'rejected': '❌ Ваш заказ отклонён.',
            'completed': '📦 Ваш заказ выполнен. Спасибо за покупку!'
        }
        
        if new_status in status_messages and old_status != new_status:
            notification = Notification(
                user_id=order.user_id,
                title=f'Заказ #{order.id} - {status_name}',
                message=f'{status_messages[new_status]} Товар: {order.product.name}, Количество: {order.quantity} {order.product.unit}',
                link=url_for('routes.my_orders')
            )
            db.session.add(notification)
            db.session.commit()
    else:
        flash('Неверный статус', 'danger')
    
    return redirect(url_for('routes.admin_orders'))


# ==================== БЫСТРОЕ ИЗМЕНЕНИЕ КОЛИЧЕСТВА ====================
@bp.route('/quick-change', methods=['POST'])
@login_required
def quick_change():
    if not current_user.can_add_product():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.products'))
    
    product_id = request.form.get('product_id')
    quantity = float(request.form.get('quantity', 1))
    action = request.args.get('type', request.form.get('type', 'plus'))
    
    product = Product.query.get_or_404(product_id)
    
    if action == 'plus':
        product.quantity += quantity
        movement_type = 'приход'
        flash_msg = f'+{quantity} {product.unit} товара "{product.name}"'
    else:
        if product.quantity >= quantity:
            product.quantity -= quantity
            movement_type = 'расход'
            flash_msg = f'-{quantity} {product.unit} товара "{product.name}"'
        else:
            flash(f'Недостаточно товара! В наличии: {product.quantity} {product.unit}', 'danger')
            return redirect(url_for('routes.products'))
    
    movement = Movement(
        product_id=product.id,
        user_id=current_user.id,
        type=movement_type,
        quantity=quantity,
        comment='Быстрое изменение через кнопки'
    )
    db.session.add(movement)
    db.session.commit()
    
    flash(flash_msg, 'success')
    return redirect(url_for('routes.products'))


# ==================== ВОССТАНОВЛЕНИЕ ПАРОЛЯ ====================
@bp.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('routes.index'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email).first()
        
        if user:
            token = secrets.token_urlsafe(32)
            session['reset_token'] = token
            session['reset_email'] = email
            session['reset_expires'] = datetime.now().timestamp() + 3600
            
            flash('Ссылка для сброса пароля отправлена на ваш email', 'info')
            return redirect(url_for('routes.reset_password', token=token))
        else:
            flash('Пользователь с таким email не найден', 'danger')
    
    return render_template('forgot_password.html')


@bp.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('routes.index'))
    
    if session.get('reset_token') != token:
        flash('Неверная или истекшая ссылка для сброса пароля', 'danger')
        return redirect(url_for('routes.forgot_password'))
    
    if datetime.now().timestamp() > session.get('reset_expires', 0):
        flash('Срок действия ссылки истек. Запросите новую.', 'danger')
        return redirect(url_for('routes.forgot_password'))
    
    if request.method == 'POST':
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if password != confirm_password:
            flash('Пароли не совпадают', 'danger')
            return redirect(url_for('routes.reset_password', token=token))
        
        if len(password) < 6:
            flash('Пароль должен содержать не менее 6 символов', 'danger')
            return redirect(url_for('routes.reset_password', token=token))
        
        user = User.query.filter_by(email=session.get('reset_email')).first()
        if user:
            user.password = bcrypt.generate_password_hash(password).decode('utf-8')
            db.session.commit()
            
            session.pop('reset_token', None)
            session.pop('reset_email', None)
            session.pop('reset_expires', None)
            
            flash('Пароль успешно изменен! Теперь вы можете войти.', 'success')
            return redirect(url_for('routes.login'))
    
    return render_template('reset_password.html')


# ==================== УВЕДОМЛЕНИЯ ====================
@bp.route('/notifications')
@login_required
def notifications():
    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).all()
    
    for notif in notifications:
        if not notif.is_read:
            notif.is_read = True
    db.session.commit()
    
    return render_template('notifications.html', notifications=notifications)


@bp.route('/notifications/count')
@login_required
def notifications_count():
    count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    return {'count': count}


@bp.route('/notification/delete/<int:notification_id>')
@login_required
def delete_notification(notification_id):
    notification = Notification.query.get_or_404(notification_id)
    
    if notification.user_id != current_user.id:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.notifications'))
    
    db.session.delete(notification)
    db.session.commit()
    flash('Уведомление удалено', 'success')
    return redirect(url_for('routes.notifications'))


@bp.route('/notifications/delete-all')
@login_required
def delete_all_notifications():
    Notification.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    flash('Все уведомления удалены', 'success')
    return redirect(url_for('routes.notifications'))